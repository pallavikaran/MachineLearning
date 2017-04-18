import openpyxl
wb = openpyxl.load_workbook('testing.xlsx')
ws = wb.get_sheet_by_name('testing_set')

import nltk

for i in range(2, 591):
	n_caps = 0
	f_essay = ws.cell(row = i, column = 3)
	essay = f_essay.value

	import re
	letters_only = re.sub("[^a-zA-Z]", " ", essay)
	essay = letters_only.lower()

	for word in essay.split():
		w = word[0]
		if w.islower():
			n_caps = n_caps + 1

	ws.cell(row = i, column = 18).value = n_caps
wb.save('testing.xlsx')