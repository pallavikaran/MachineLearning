from nltk.stem.lancaster import LancasterStemmer
from nltk.stem import WordNetLemmatizer
lancaster_stemmer = LancasterStemmer()
wordnet_lemmatizer = WordNetLemmatizer()

import openpyxl
wb1 = openpyxl.load_workbook('training.xlsx')
ws1 = wb1.get_sheet_by_name('training_set')

f_essay = ws1.cell(row = 17, column = 3)
essay1 = f_essay.value

f_essay = ws1.cell(row = 226, column = 3)
essay2 = f_essay.value

f_essay = ws1.cell(row = 410, column = 3)
essay3 = f_essay.value

f_essay = ws1.cell(row = 816, column = 3)
essay4 = f_essay.value

f_essay = ws1.cell(row = 1131, column = 3)
essay5 = f_essay.value

sample_essay = essay1 + essay2 + essay3 + essay4 + essay5

import re
letters_only = re.sub("[^a-zA-Z]", " ", sample_essay)

lower_case = letters_only.lower()
words = lower_case.split()

from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords

word_bank = [w for w in words if not w in stopwords.words("english")]
stem = [lancaster_stemmer.stem(w) for w in word_bank]
word_bank = [wordnet_lemmatizer.lemmatize(w) for w in stem]

wb2 = openpyxl.load_workbook('testing.xlsx')
ws2 = wb2.get_sheet_by_name('testing_set')

for i in range(2, 591):
	count = 0
	a = 0.0
	b = 0.0
	c = 0.0
	
	f_essay = ws2.cell(row = i, column = 3)
	essay = f_essay.value
	
	letters_only = re.sub("[^a-zA-Z]", " ", essay)
	lower_case = letters_only.lower()
	words = lower_case.split()
	local_bank = [w for w in words if not w in stopwords.words("english")]
	stem = [lancaster_stemmer.stem(w) for w in local_bank]
	local_bank = [wordnet_lemmatizer.lemmatize(w) for w in stem]
	
	count = set(word_bank).intersection(local_bank)
	a = len(count) * 100
	b = len(word_bank)
	c = a / b
	
	ws2.cell(row = i, column = 15).value = c
wb2.save('testing.xlsx')

