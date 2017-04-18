import openpyxl
wb = openpyxl.load_workbook('testing.xlsx')
ws = wb.get_sheet_by_name('testing_set')

import enchant
dictionary = enchant.Dict("en_US")
 
for i in range(2, 591):
	temp = 0
	f_essay = ws.cell(row = i, column = 3)
	essay = f_essay.value
	
	import re
	letters_only = re.sub("[^a-zA-Z]", " ", essay)
	essay = letters_only.lower()
	
	for word in essay.split():
		if dictionary.check(word) == 0:
			temp = temp + 1
	ws.cell(row = i, column = 13).value = temp
wb.save('testing.xlsx')
	
	

