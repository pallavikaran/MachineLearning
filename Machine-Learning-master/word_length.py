import openpyxl
wb = openpyxl.load_workbook('testing.xlsx')
ws = wb.get_sheet_by_name('testing_set')

import nltk

for i in range(2, 591):
	word_5 = 0
	word_4 = 0
	f_essay = ws.cell(row = i, column = 3)
	essay = f_essay.value

	import re
	letters_only = re.sub("[^a-zA-Z]", " ", essay)
	essay = letters_only.lower()

	for word in essay.split():
		if len(word) >= 5:
			word_5 = word_5 + 1
		if len(word) <= 4:
			word_4 = word_4 + 1

	ws.cell(row = i, column = 16).value = word_4
	ws.cell(row = i, column = 17).value = word_5
wb.save('testing.xlsx')