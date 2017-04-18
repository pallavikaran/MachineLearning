import openpyxl
wb = openpyxl.load_workbook('testing.xlsx')
ws = wb.get_sheet_by_name('testing_set')

for i in range(2, 590):
	score = ws.cell(row = i, column = 11)
	score1 = score.value
	if score1 == 0:
		score1 = 0
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 1:
		score1 = 1
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 2:
		score1 = 2
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 3:
		score1 = 3
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 4:
		score1 = 3
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 5:
		score1 = 4
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 6:
		score1 = 5
		ws.cell(row = i, column = 11).value = score1

	elif score1 == 7:
		score1 = 6
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 8:
		score1 = 7
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 9:
		score1 = 8
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 10:
		score1 = 8
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 11:
		score1 = 9
		ws.cell(row = i, column = 11).value = score1
	
	elif score1 == 12:
		score1 = 10
		ws.cell(row = i, column = 11).value = score1

wb.save('testing.xlsx')