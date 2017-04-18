import openpyxl
wb = openpyxl.load_workbook('training.xlsx')
ws = wb.get_sheet_by_name('training_set')

for i in range(2, 1785):
	score = ws.cell(row = i, column = 4)
	score1 = score.value
	if score1 == 0:
		score1 = 0
		ws.cell(row = i, column = 4).value = score1
	
	elif score1 == 1:
		score1 = 1.67
		ws.cell(row = i, column = 4).value = score1
	
	elif score1 == 2:
		score1 = 3.34
		ws.cell(row = i, column = 4).value = score1
	
	elif score1 == 3:
		score1 = 5
		ws.cell(row = i, column = 4).value = score1
	
	elif score1 == 4:
		score1 = 6.67
		ws.cell(row = i, column = 4).value = score1
	
	elif score1 == 5:
		score1 = 8.34
		ws.cell(row = i, column = 4).value = score1
	
	elif score1 == 6:
		score1 = 10
		ws.cell(row = i, column = 4).value = score1

for i in range(2, 1785):
	score = ws.cell(row = i, column = 5)
	score2 = score.value
	if score2 == 0:
		score2 = 0
		ws.cell(row = i, column = 5).value = score2
	
	elif score2 == 1:
		score2 = 1.67
		ws.cell(row = i, column = 5).value = score2
	
	elif score2 == 2:
		score2 = 3.34
		ws.cell(row = i, column = 5).value = score2
	
	elif score2 == 3:
		score2 = 5
		ws.cell(row = i, column = 5).value = score2
	
	elif score2 == 4:
		score2 = 6.67
		ws.cell(row = i, column = 5).value = score2
	
	elif score2 == 5:
		score2 = 8.34
		ws.cell(row = i, column = 5).value = score2
	
	elif score2 == 6:
		score2 = 10
		ws.cell(row = i, column = 5).value = score2

for i in range(2, 1785):
	score = ws.cell(row = i, column = 4)
	score1 = score.value
	score = ws.cell(row = i, column = 5)
	score2 = score.value
	
	avg_score = (score1 + score2)/2
	ws.cell(row = i, column = 6).value = avg_score

wb.save('training.xlsx')