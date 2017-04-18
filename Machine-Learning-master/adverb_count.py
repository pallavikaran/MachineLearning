import openpyxl
wb = openpyxl.load_workbook('training.xlsx')
ws = wb.get_sheet_by_name('training_set')

import nltk
from nltk import word_tokenize, pos_tag

for i in range(2, 1785):
	count = 0
	f_essay = ws.cell(row = i, column = 3)
	essay = f_essay.value

	import re
	letters_only = re.sub("[^a-zA-Z]", " ", essay)
	essay = letters_only.lower()

	x = nltk.word_tokenize(essay)
	y = nltk.pos_tag(x)
	nouns = [word for word, pos in y if pos == 'RB']
	count = len(nouns)
	ws.cell(row = i, column = 12).value = count
wb.save('training.xlsx')