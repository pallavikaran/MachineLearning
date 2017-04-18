import openpyxl
wb = openpyxl.load_workbook('testing.xlsx')
ws = wb.get_sheet_by_name('testing_set')

import textstat
from textstat.textstat import textstat

for i in range(2, 591):
	score = 0
	f_essay = ws.cell(row = i, column = 3)
	essay = f_essay.value
	score = textstat.flesch_reading_ease(essay)
	ws.cell(row = i, column = 11).value = score
wb.save('testing.xlsx')