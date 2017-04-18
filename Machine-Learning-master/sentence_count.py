import openpyxl
wb = openpyxl.load_workbook('testing.xlsx')
ws = wb.get_sheet_by_name('testing_set')

import nltk

for i in range(2, 591):
	count = 0
	f_essay = ws.cell(row = i, column = 3)
	essay = f_essay.value
	count = essay.count(".")
	ws.cell(row = i, column = 8).value = count
wb.save('testing.xlsx')